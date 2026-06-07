import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

def generate_volunteers_excel(volunteers):
    """
    Generate an Excel file containing volunteer data.
    
    Args:
        volunteers: List of Volunteer model objects.
    
    Returns:
        str: File path to the generated Excel file.
    """
    # Ensure the uploads directory exists
    os.makedirs("uploads", exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Volunteers"
    
    # Define headers
    headers = [
        "ID",
        "Registration No",
        "Name",
        "Phone",
        "Gender",
        "Age",
        "LGA",
        "Ward",
        "Unit",
        "Highest Qualification",
        "Employment Status"
    ]
    
    # Style the header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center")
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Append volunteer data
    for volunteer in volunteers:
        ws.append([
            volunteer.id,
            volunteer.registration_no,
            volunteer.name,
            volunteer.phone,
            volunteer.gender,
            volunteer.age,
            volunteer.lga,
            volunteer.ward,
            volunteer.unit,
            volunteer.highest_qualification,
            volunteer.employment_status
        ])
    
    # Adjust column widths for readability
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Generate a unique filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path = f"uploads/volunteers_{timestamp}.xlsx"
    
    wb.save(file_path)
    return file_path